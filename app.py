import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import tempfile
from openpyxl.utils.exceptions import InvalidFileException
from copy import copy
import base64
import io
import traceback

# アプリのタイトルとスタイルの設定
st.set_page_config(page_title="Excelシートマージツール（マクロ対応版）", layout="wide")

# アプリのタイトルと説明
st.title("Excelシートマージツール（マクロ対応版）")
st.markdown("""
### このツールの機能
- マクロ付きのExcelテンプレートと追加データのExcelファイルを受け取ります
- テンプレートの最初の2つのシートはそのまま保持します
- 追加データの全シートをSheet1, Sheet2...という名前でテンプレートに追加します
- マクロは保持されます（出力は.xlsm形式）
""")

# ファイルのダウンロード用関数
def get_binary_file_downloader_html(bin_data, file_label='File', file_name='file.xlsm'):
    bin_str = base64.b64encode(bin_data).decode()
    href = f'<a href="data:application/vnd.ms-excel.sheet.macroEnabled.12;base64,{bin_str}" download="{file_name}">📥 {file_label}</a>'
    return href

def load_template(file_content):
    """テンプレートExcelファイル(マクロ有り)を読み込む関数"""
    try:
        # io.BytesIOを使用してファイルコンテンツをメモリ上で扱う
        file_stream = io.BytesIO(file_content)
        
        # keep_vba=True でVBAマクロを保持
        template_wb = load_workbook(file_stream, keep_vba=True)
        
        # テンプレートに2つのシートがあることを確認
        if len(template_wb.sheetnames) < 2:
            st.error("エラー: テンプレートには最低2つのシートが必要です")
            return None
            
        return template_wb
    except Exception as e:
        st.error(f"テンプレートファイルの読み込みエラー: {str(e)}")
        st.error(traceback.format_exc())
        return None

def load_additional_data(file_content):
    """追加データのExcelファイルを読み込む関数"""
    try:
        # io.BytesIOを使用してファイルコンテンツをメモリ上で扱う
        file_stream = io.BytesIO(file_content)
        
        # VBAコードがなくても問題ないが、あれば保持
        data_wb = load_workbook(file_stream, keep_vba=True)
        return data_wb
    except Exception as e:
        st.error(f"追加データファイルの読み込みエラー: {str(e)}")
        st.error(traceback.format_exc())
        return None

def merge_workbooks(template_wb, data_wb):
    """
    データワークブックのシートをテンプレートワークブックに追加する関数
    テンプレートのシートはそのままで、追加シートは「Sheet1」から連番で命名
    """
    try:
        # 元のテンプレートシート名を表示
        template_sheet_names = template_wb.sheetnames.copy()
        st.write(f"既存テンプレートシート（維持）: {', '.join(template_sheet_names)}")
        
        # シート名のマッピングを保存するためのディクショナリ（元のシート名→新しいシート名）
        sheet_name_mapping = {}
        
        # 追加シートの番号は1から開始
        sheet_number = 1
        
        # 進捗バーを表示
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # 追加データのシートをコピー
        total_sheets = len(data_wb.sheetnames)
        for idx, sheet_name in enumerate(data_wb.sheetnames):
            # 進捗を更新
            progress = int((idx / total_sheets) * 100)
            progress_bar.progress(progress)
            status_text.text(f"処理中... {sheet_name} ({idx+1}/{total_sheets})")
            
            # データのシートを取得
            source_sheet = data_wb[sheet_name]
            
            # 新しいシート名を生成（Sheet1, Sheet2, ...）
            new_sheet_name = f"Sheet{sheet_number}"
            sheet_number += 1
            
            # 同名のシートがある場合は番号を調整
            while new_sheet_name in template_wb.sheetnames:
                new_sheet_name = f"Sheet{sheet_number}"
                sheet_number += 1
            
            # マッピングを保存
            sheet_name_mapping[sheet_name] = new_sheet_name
            
            # テンプレートに新しいシートを作成
            target_sheet = template_wb.create_sheet(title=new_sheet_name)
            
            # セルの幅と高さをコピー
            for i, col in enumerate(source_sheet.columns, 1):
                col_letter = openpyxl.utils.get_column_letter(i)
                if col_letter in source_sheet.column_dimensions:
                    target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
            
            for i, row in enumerate(source_sheet.rows, 1):
                if i in source_sheet.row_dimensions:
                    target_sheet.row_dimensions[i].height = source_sheet.row_dimensions[i].height
            
            # まず、全セルの値とスタイルをコピー（結合セルの処理前）
            for row in source_sheet.rows:
                for cell in row:
                    # 結合セルでないセルのみコピー
                    if isinstance(cell, openpyxl.cell.cell.Cell):  # MergedCellでなくCellの場合のみ処理
                        target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                        target_cell.value = cell.value
                        
                        # スタイルをコピー
                        if cell.has_style:
                            target_cell.font = copy(cell.font)
                            target_cell.border = copy(cell.border)
                            target_cell.fill = copy(cell.fill)
                            target_cell.number_format = cell.number_format
                            target_cell.protection = copy(cell.protection)
                            target_cell.alignment = copy(cell.alignment)
            
            # 結合セルの情報をコピー（値とスタイルをコピーした後）
            for merged_range in source_sheet.merged_cells.ranges:
                # 同じ範囲をターゲットシートでも結合
                target_sheet.merge_cells(str(merged_range))
        
        # 進捗バーを完了
        progress_bar.progress(100)
        status_text.text("処理完了！")
        
        # シート名のマッピング情報を表示
        st.subheader("元のシート名と新しいシート名のマッピング:")
        mapping_data = {"元のシート名": list(sheet_name_mapping.keys()), 
                         "新しいシート名": list(sheet_name_mapping.values())}
        st.table(pd.DataFrame(mapping_data))
        
        return template_wb
    
    except Exception as e:
        st.error(f"ワークブックのマージエラー: {str(e)}")
        st.error(traceback.format_exc())
        return None

def main():
    # サイドバーに説明文を表示
    st.sidebar.title("使い方")
    st.sidebar.markdown("""
    1. マクロ付きテンプレートファイルをアップロード (.xlsm推奨)
    2. 追加データExcelファイルをアップロード
    3. 「処理開始」ボタンをクリック
    4. 処理完了後、結果ファイルをダウンロード
    
    **注意点:**
    - テンプレートには最低2つのシートが必要です
    - マクロは保持されます
    - 出力ファイルは.xlsm形式です
    """)
    
    # デバッグ情報表示（問題解決後に削除可能）
    st.sidebar.markdown("---")
    st.sidebar.subheader("デバッグ情報")
    st.sidebar.info("アプリが正常に読み込まれました")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("1. マクロ付きテンプレートファイル")
        template_file = st.file_uploader("テンプレートファイルをアップロード", type=["xlsm", "xlsx"])
    
    with col2:
        st.subheader("2. 追加データファイル")
        data_file = st.file_uploader("追加データファイルをアップロード", type=["xlsx", "xlsm", "xls"])
    
    st.markdown("---")
    
    if template_file is not None and data_file is not None:
        st.info("ファイルがアップロードされました。「処理開始」ボタンをクリックしてください。")
        
        # プロセス開始ボタン
        if st.button("処理開始", key="process_button", help="クリックしてファイルをマージします"):
            st.info("処理を開始します。しばらくお待ちください...")
            
            try:
                # ファイルコンテンツを直接読み込む
                template_content = template_file.read()
                data_content = data_file.read()
                
                # テンプレートを読み込み
                template_wb = load_template(template_content)
                if template_wb is not None:
                    # 追加データを読み込み
                    data_wb = load_additional_data(data_content)
                    if data_wb is not None:
                        # ワークブックをマージ
                        merged_wb = merge_workbooks(template_wb, data_wb)
                        if merged_wb is not None:
                            try:
                                # マージされたファイルをメモリ上に保存
                                output_buffer = io.BytesIO()
                                merged_wb.save(output_buffer)
                                output_buffer.seek(0)
                                
                                st.success("マージが完了しました！")
                                
                                # ダウンロードリンクを表示
                                output_filename = "merged_excel.xlsm"
                                st.markdown(
                                    get_binary_file_downloader_html(
                                        output_buffer.getvalue(), 
                                        "結果ファイルをダウンロード", 
                                        output_filename
                                    ), 
                                    unsafe_allow_html=True
                                )
                                
                                # 処理の概要を表示
                                st.subheader("処理の概要")
                                st.markdown(f"""
                                - テンプレートの元のシート数: {len(template_wb.sheetnames) - len(data_wb.sheetnames)}
                                - 追加されたシート数: {len(data_wb.sheetnames)}
                                - 最終シート数: {len(merged_wb.sheetnames)}
                                """)
                            except Exception as e:
                                st.error(f"ファイル保存エラー: {str(e)}")
                                st.error(traceback.format_exc())
            except Exception as e:
                st.error(f"予期せぬエラーが発生しました: {str(e)}")
                st.error(traceback.format_exc())
    else:
        st.info("テンプレートファイルと追加データファイルをアップロードしてください。")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        st.error(f"アプリケーションエラー: {str(e)}")
        st.error(traceback.format_exc())
